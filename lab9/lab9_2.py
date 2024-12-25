from openpyxl import Workbook, load_workbook
from openpyxl.chart import PieChart, Reference

file_path = 'lab9/table.xlsx'
wb = load_workbook(file_path)
ws = wb.active

ws["A14"] = "Средняя зарплата по отделам:"
ws["B14"] = (ws[4][5].value + ws[9][5].value + ws[11][5].value) / 3
minimum = ws[2][5].value
maximum = 0
for i in range(2, 10):
    if i not in [4, 9]:
        minimum = min(minimum, ws[i][5].value)
        maximum = max(maximum, ws[i][5].value)
ws["A15"] = "Максимальная зарплата:"
ws["B15"] = maximum
ws["A16"] = "Минимальная зарплата:"
ws["B16"] = minimum

wb.save("lab9/table.xlsx")